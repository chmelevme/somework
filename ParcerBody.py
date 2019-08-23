from bs4 import BeautifulStoneSoup
import requests
import openpyxl
from multiprocessing import Pool


class Some_Exeption(Exception):
    pass


class Connection_Exception(Some_Exeption):
    pass


class Format_Exeption(Some_Exeption):
    pass


class SuperParser():
    """Класс содержащий базовые инструменты"""

    def __init__(self, headers, urls):
        self.headers = headers
        self.urls = urls
        self.session = requests.session()

    def get_all_urls(self):
        """Возвращает список url"""
        list_of_urls = []
        for url in self.urls:
            request = self.session.get(url, headers=self.headers)
            soap = BeautifulStoneSoup(request.content)
            urls = soap.find_all('loc')
            list_of_urls += [url.next_element for url in urls]
        return list_of_urls

    def parse_and_load(self, url):

        try:
            final = self.parse_data(url)
        except Some_Exeption:
            print("Ощибка в соединении или имени")
            return
        print(final, end='\n\n')
        return final

    def parse_and_save(self, patch):
        patch = patch.replace('\ ', ' \\ ')

        vk = openpyxl.Workbook()
        sh = vk.active
        sh.title = "Цены"
        sh.cell(column=1, row=1).value = "Название"
        sh.cell(column=2, row=1).value = "Номер"
        sh.cell(column=3, row=1).value = "Оптовая цена"
        sh.cell(column=4, row=1).value = "Цена розничного магазина"
        sh.cell(column=5, row=1).value = "Ссылка на товар"
        vk.save(patch)

        list_of_urls = self.get_all_urls()
        len_list_of_urls = len(list_of_urls)
        vk = openpyxl.load_workbook(patch)
        sh = vk.active
        sh.cell(column=6, row=1).value = len_list_of_urls
        for idx, item in enumerate(list_of_urls):
            idx += 2
            result = self.parse_and_load(item)
            try:
                sh.cell(column=1, row=idx).value = str(result['name_of_product'])
                sh.cell(column=2, row=idx).value = str(result['reference'])
                sh.cell(column=3, row=idx).value = str(result['price_for_all'])
                sh.cell(column=4, row=idx).value = str(result['price_for_registered'])
                sh.cell(column=5, row=idx).value = str(result['url'])
            except TypeError:
                pass

        if idx % 1000 == 0:
            sh.cell(column=6, row=idx).value = idx / len_list_of_urls * 100
            vk.save(patch)
            vk = openpyxl.load_workbook(patch)
            sh = vk.active
            vk.save(patch)


def main():
    print('Good')


if __name__ == "__main__":
    main()
